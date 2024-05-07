import datetime
import os
import urllib.request

import openpyxl
from bs4 import BeautifulSoup
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border
from openpyxl.styles.borders import Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.ui import WebDriverWait


# ITmediaNEWSでカテゴリが速報のNEWSを一覧化して出力するツール

def exist_ws(target_name, work_book):
    check = False

    # ファイル内の全てのシートをループして検索
    for work_sheet in work_book.worksheets:

        # 指定シートが存在していれば、変数にTrueを格納
        if work_sheet.title == target_name:
            check = True

    return check


def move_sheet(work_book, sheet_name_from):
    # シートの並び替え（日付降順）
    for loop_count_sheet, worksheet_to in enumerate(work_book.worksheets):
        # シート名の取得
        sheet_name_to = worksheet_to.title

        # 比較先シート名をdate型に変換
        date_to = datetime.datetime.strptime(sheet_name_to, '%Y年%m月%d日')
        # 比較元シート名をdate型に変換
        date_from = datetime.datetime.strptime(sheet_name_from, '%Y年%m月%d日')
        # 作成したシートが既存のシートよりも日付が後の場合
        if date_to < date_from:

            for sheet_index, worksheet_from in enumerate(work_book.worksheets):

                if sheet_name_from in worksheet_from.title:
                    # 移動先までの移動シート枚数
                    sheet_num = sheet_index - loop_count_sheet
                    # シート移動
                    wb.move_sheet(worksheet_from, offset=-sheet_num)
                    break
            break


driver = webdriver.Chrome()

# URLを設定する
URL = "https://www.itmedia.co.jp/news/"

# 最大の読み込み時間を設定 今回は最大30秒待機できるようにする
wait = WebDriverWait(driver=driver, timeout=30)

# 検索カテゴリー
search = "速報"

# 結果出力ファイルパス
EXCEL_FILE_PATH = "output/output.xlsx"

# ファイル存在フラグ
file_exist_flag = False

# Excel Workbookオブジェクトを取得
if not os.path.exists(EXCEL_FILE_PATH):
    wb = openpyxl.Workbook()
else:
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH)
    file_exist_flag = True

try:
    # サイトに移動
    driver.get(URL)

    categories = driver.find_elements(By.CLASS_NAME, "g_lnav_o")

    for category in categories:
        # 対象のカテゴリーの場合はリンク先へ遷移する
        if category.text in search:
            category.find_element(By.TAG_NAME, "a").click()
            break
    else:
        print("検索対象のカテゴリーが存在しませんでした")
        raise Exception

    wait.until(ec.presence_of_all_elements_located)

    # 速報情報を日付毎に配列として取得
    days_topic = driver.find_elements(By.CLASS_NAME, "colBox.colBoxToday")

    # href_list = []

    # 日付分繰り返す
    for loop_count_day, day_topic in enumerate(days_topic):

        # 日付の取得
        day = day_topic.find_element(By.TAG_NAME, "h2").text

        if file_exist_flag:
            if exist_ws(day, wb):
                ws = wb[day]
            else:
                ws = wb.create_sheet(day)
                # シートの並び替え（日付降順）
                move_sheet(wb, day)

        else:
            ws = wb.create_sheet(day)

        # 列の幅の設定
        ws.column_dimensions["A"].width = 18.5
        ws.column_dimensions["B"].width = 79.38
        ws.column_dimensions["C"].width = 79.38
        ws.column_dimensions["D"].width = 59

        # ヘッダ書式の設定
        pf = PatternFill(patternType="solid", fgColor="ffe4b5")
        font = Font(name="メイリオ", bold=True)
        side1 = Side(style="thin", color="000000")  # 細線（黒）
        side2 = Side(style="double", color="000000")  # 細線（二重線）
        border = Border(top=side1, bottom=side2, left=side1, right=side1)
        cells_all = ws["A1":"D1"]
        for cells in cells_all:
            for cell in cells:
                cell.fill = pf
                cell.font = font
                cell.border = border

        # ヘッダ名の設定
        ws.cell(column=1, row=1).value = "公開日"
        ws.cell(column=2, row=1).value = "タイトル"
        ws.cell(column=3, row=1).value = "内容"
        ws.cell(column=4, row=1).value = "リンク先"

        # 速報情報
        topics = day_topic.find_elements(By.TAG_NAME, "li")

        # 中身の書式の設定(フォント)
        font = Font(name="メイリオ")
        # 中身の書式の設定(崖線)
        border = Border(bottom=side1, left=side1, right=side1)

        # 速報記事分繰り返す
        for loop_count_topic, topic in enumerate(topics):

            # 書式設定
            cells_all = ws["A{}".format(2 + loop_count_topic):"D{}".format(2 + loop_count_topic)]
            for cells in cells_all:
                for cell in cells:
                    cell.font = font
                    cell.border = border

            date = topic.find_element(By.CLASS_NAME, "colBoxUlistDate").text
            title = topic.find_element(By.TAG_NAME, "a").get_attribute("title")
            href = topic.find_element(By.TAG_NAME, "a").get_attribute("href")

            ws.cell(column=1, row=2 + loop_count_topic).value = topic.find_element(By.CLASS_NAME,
                                                                                   "colBoxUlistDate").text  # 日付
            ws.cell(column=2, row=2 + loop_count_topic).value = title  # タイトル
            ws.cell(column=4, row=2 + loop_count_topic).value = href  # リンク先

            html = urllib.request.urlopen(href)

            soup = BeautifulSoup(html, "html.parser")
            ps = soup.find_all("p")
            content = ""
            for p in ps:
                content += p.text + "\n"
            ws.cell(column=3, row=2 + loop_count_topic).value = content  # 内容

    if exist_ws("Sheet", wb):
        del wb["Sheet"]

    for sheet in wb:
        wb[sheet.title].views.sheetView[0].tabSelected = False
    # Excelファイルを保存
    wb.save(EXCEL_FILE_PATH)


except Exception as e:
    print(e)
    print("エラーが発生しました。")

# 最後にドライバーを終了する
finally:
    driver.quit()
    print("終了")
